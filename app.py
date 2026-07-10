import os
import io
import psycopg2
from psycopg2.extras import RealDictCursor
from flask import Flask, render_template, request, redirect, session, flash, url_for, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
import cloudinary
import cloudinary.uploader

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'tanuku-drains-2026-secret')

DATABASE_URL = os.environ.get('DATABASE_URL')

cloudinary.config(
    cloud_name = os.environ.get('CLOUDINARY_CLOUD_NAME'),
    api_key = os.environ.get('CLOUDINARY_API_KEY'),
    api_secret = os.environ.get('CLOUDINARY_API_SECRET')
)

def get_db_connection():
    if not DATABASE_URL:
        raise Exception("DATABASE_URL not set")
    conn = psycopg2.connect(DATABASE_URL)
    return conn

def init_db():
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id SERIAL PRIMARY KEY,
                username VARCHAR(50) UNIQUE NOT NULL,
                password_hash VARCHAR(255) NOT NULL,
                role VARCHAR(20) DEFAULT 'user',
                ward VARCHAR(10)
            )
        ''')

        cur.execute('''
            CREATE TABLE IF NOT EXISTS drains (
                id SERIAL PRIMARY KEY,
                drain_id VARCHAR(50),
                ward VARCHAR(10) NOT NULL,
                location TEXT,
                status VARCHAR(50) DEFAULT 'Pending',
                work_type VARCHAR(100),
                work_date DATE,
                updated_by VARCHAR(50),
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        cur.execute("""
            DO $$
            BEGIN
                IF NOT EXISTS (
                    SELECT 1 FROM pg_constraint
                    WHERE conname = 'drains_drain_id_ward_key'
                ) THEN
                    ALTER TABLE drains ADD CONSTRAINT drains_drain_id_ward_key UNIQUE (drain_id, ward);
                END IF;
            END $$;
        """)

        cur.execute('''
            CREATE TABLE IF NOT EXISTS drain_work_logs (
                id SERIAL PRIMARY KEY,
                drain_id INTEGER REFERENCES drains(id) ON DELETE CASCADE,
                photo_url TEXT NOT NULL,
                work_type VARCHAR(100) NOT NULL,
                status VARCHAR(50) NOT NULL,
                remarks TEXT,
                uploaded_by VARCHAR(50) NOT NULL,
                uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        cur.execute("""
            INSERT INTO users (username, password_hash, role, ward)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (username) DO NOTHING
        """, ('admin', generate_password_hash('Tanuku@2026'), 'admin', None))

        ward_users = [
            ('sanitation11', 'Sanitation11@2026', '11'),
            ('sanitation12', 'Sanitation12@2026', '12'),
            ('sanitation29', 'Sanitation29@2026', '29')
        ]

        for username, password, ward in ward_users:
            cur.execute("""
                INSERT INTO users (username, password_hash, role, ward)
                VALUES (%s, %s, %s, %s)
                ON CONFLICT (username) DO NOTHING
            """, (username, generate_password_hash(password), 'user', ward))

        conn.commit()
        cur.close()
        conn.close()
        print("Database initialized successfully")
    except Exception as e:
        print(f"Database init error: {e}")

try:
    init_db()
except Exception as e:
    print(f"Database init error: {e}")

@app.route('/')
def index():
    if 'username' in session:
        return redirect('/dashboard')
    return redirect('/login')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("SELECT * FROM users WHERE username = %s", (username,))
        user = cur.fetchone()
        cur.close()
        conn.close()

        if user and check_password_hash(user['password_hash'], password):
            session['username'] = user['username']
            session['role'] = user['role']
            session['ward'] = user['ward']
            return redirect('/dashboard')
        else:
            return render_template('login.html', error='Invalid username or password')

    return render_template('login.html')

@app.route('/dashboard')
def dashboard():
    if 'username' not in session:
        return redirect('/login')

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    if session.get('role') == 'admin':
        cur.execute("""
            SELECT d.*,
                   (SELECT COUNT(*) FROM drain_work_logs WHERE drain_id = d.id) as work_count,
                   (SELECT uploaded_at FROM drain_work_logs WHERE drain_id = d.id ORDER BY uploaded_at DESC LIMIT 1) as last_work_date
            FROM drains d
            ORDER BY ward, drain_id
        """)
    else:
        cur.execute("""
            SELECT d.*,
                   (SELECT COUNT(*) FROM drain_work_logs WHERE drain_id = d.id) as work_count,
                   (SELECT uploaded_at FROM drain_work_logs WHERE drain_id = d.id ORDER BY uploaded_at DESC LIMIT 1) as last_work_date
            FROM drains d
            WHERE ward = %s
            ORDER BY drain_id
        """, (session.get('ward'),))

    drains = cur.fetchall()
    cur.close()
    conn.close()

    return render_template('dashboard.html', drains=drains)

@app.route('/upload_work/<int:drain_id>', methods=['GET', 'POST'])
def upload_work(drain_id):
    if 'username' not in session:
        return redirect('/login')

    if session.get('role') == 'admin':
        flash('Admin cannot upload work photos. Only ward users can upload.')
        return redirect('/dashboard')

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    cur.execute("SELECT * FROM drains WHERE id = %s", (drain_id,))
    drain = cur.fetchone()

    if not drain:
        flash('Drain not found')
        cur.close()
        conn.close()
        return redirect('/dashboard')

    if drain['ward']!= session.get('ward'):
        flash('You do not have access to this drain')
        cur.close()
        conn.close()
        return redirect('/dashboard')

    cur.execute("""
        SELECT * FROM drain_work_logs
        WHERE drain_id = %s
        ORDER BY uploaded_at DESC
    """, (drain_id,))
    work_history = cur.fetchall()

    if request.method == 'POST':
        work_type = request.form.get('work_type')
        status = request.form.get('status')
        remarks = request.form.get('remarks', '')

        if 'photo' not in request.files or request.files['photo'].filename == '':
            flash('Photo is required for work upload')
            cur.close()
            conn.close()
            return redirect(url_for('upload_work', drain_id=drain_id))

        photo = request.files['photo']
        try:
            upload_result = cloudinary.uploader.upload(photo, folder="tanuku_drains")
            photo_url = upload_result['secure_url']
        except Exception as e:
            flash(f'Photo upload failed: {str(e)}')
            cur.close()
            conn.close()
            return redirect(url_for('upload_work', drain_id=drain_id))

        cur.execute("""
            INSERT INTO drain_work_logs (drain_id, photo_url, work_type, status, remarks, uploaded_by)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (drain_id, photo_url, work_type, status, remarks, session['username']))

        cur.execute("""
            UPDATE drains
            SET status = %s, work_type = %s, work_date = CURRENT_DATE,
                updated_by = %s, updated_at = CURRENT_TIMESTAMP
            WHERE id = %s
        """, (status, work_type, session['username'], drain_id))

        conn.commit()
        cur.close()
        conn.close()
        flash('Work data uploaded successfully')
        return redirect('/dashboard')

    cur.close()
    conn.close()
    return render_template('upload_work.html', drain=drain, work_history=work_history)

@app.route('/import_excel', methods=['GET', 'POST'])
def import_excel():
    if 'username' not in session:
        return redirect('/login')

    if session.get('role')!= 'admin':
        flash('Only admin can import drain master data')
        return redirect('/dashboard')

    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file selected')
            return redirect('/import_excel')

        file = request.files['file']

        if file.filename == '':
            flash('No file selected')
            return redirect('/import_excel')

        if file and file.filename.endswith('.xlsx'):
            try:
                wb = openpyxl.load_workbook(file)
                sheet = wb.active

                conn = get_db_connection()
                cur = conn.cursor()

                count = 0
                skipped = 0
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    drain_id = str(row[2]) if len(row) > 2 and row[2] else ''
                    ward = str(row[1]) if len(row) > 1 and row[1] else ''
                    location = str(row[3]) if len(row) > 3 and row[3] else ''

                    if not drain_id or not ward:
                        skipped += 1
                        continue

                    cur.execute("""
                        INSERT INTO drains (drain_id, ward, location, status, updated_by)
                        VALUES (%s, %s, %s, 'Pending', %s)
                        ON CONFLICT (drain_id, ward)
                        DO UPDATE SET location = EXCLUDED.location, updated_by = EXCLUDED.updated_by, updated_at = CURRENT_TIMESTAMP
                    """, (drain_id, ward, location, session['username']))
                    count += 1

                cur.execute("SELECT DISTINCT ward FROM drains")
                all_wards = [row[0] for row in cur.fetchall()]
                for ward in all_wards:
                    if ward:
                        username = f'sanitation{ward}'
                        password = f'Sanitation{ward}@2026'
                        cur.execute("""
                            INSERT INTO users (username, password_hash, role, ward)
                            VALUES (%s, %s, %s, %s)
                            ON CONFLICT (username) DO NOTHING
                        """, (username, generate_password_hash(password), 'user', ward))

                conn.commit()
                cur.close()
                conn.close()

                flash(f'Successfully imported {count} drains. Skipped {skipped} rows. Data reflected in ward logins.')
                return redirect('/dashboard')

            except Exception as e:
                flash(f'Error importing Excel: {str(e)}')
                return redirect('/import_excel')
        else:
            flash('Please upload a.xlsx file')
            return redirect('/import_excel')

    return render_template('import_excel.html')

# NEW: Delete single drain - Admin only
@app.route('/delete_drain/<int:drain_id>', methods=['POST'])
def delete_drain(drain_id):
    if 'username' not in session or session.get('role')!= 'admin':
        flash('Only admin can delete drains')
        return redirect('/dashboard')

    conn = get_db_connection()
    cur = conn.cursor()

    # Get drain info for confirmation message
    cur.execute("SELECT drain_id, ward FROM drains WHERE id = %s", (drain_id,))
    drain = cur.fetchone()

    if drain:
        cur.execute("DELETE FROM drains WHERE id = %s", (drain_id,))
        conn.commit()
        flash(f'Drain {drain[0]} from Ward {drain[1]} deleted successfully. All work logs also deleted.')
    else:
        flash('Drain not found')

    cur.close()
    conn.close()
    return redirect('/dashboard')

# NEW: Delete all drains in a ward - Admin only
@app.route('/delete_ward_data/<ward>', methods=['POST'])
def delete_ward_data(ward):
    if 'username' not in session or session.get('role')!= 'admin':
        flash('Only admin can delete ward data')
        return redirect('/dashboard')

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT COUNT(*) FROM drains WHERE ward = %s", (ward,))
    count = cur.fetchone()[0]

    if count > 0:
        cur.execute("DELETE FROM drains WHERE ward = %s", (ward,))
        conn.commit()
        flash(f'All {count} drains and work logs from Ward {ward} deleted successfully.')
    else:
        flash(f'No drains found in Ward {ward}')

    cur.close()
    conn.close()
    return redirect('/dashboard')

# NEW: Delete all work logs only - keeps drain master data
@app.route('/delete_all_work_logs', methods=['POST'])
def delete_all_work_logs():
    if 'username' not in session or session.get('role')!= 'admin':
        flash('Only admin can delete work logs')
        return redirect('/dashboard')

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT COUNT(*) FROM drain_work_logs")
    count = cur.fetchone()[0]

    cur.execute("DELETE FROM drain_work_logs")

    # Reset drain status to Pending
    cur.execute("""
        UPDATE drains SET status = 'Pending', work_type = NULL, work_date = NULL, updated_by = %s, updated_at = CURRENT_TIMESTAMP
    """, (session['username'],))

    conn.commit()
    cur.close()
    conn.close()
    flash(f'All {count} work logs deleted. Drain master data retained. All drain status reset to Pending.')
    return redirect('/dashboard')

# NEW: Clear entire database - Admin only - DANGER
@app.route('/clear_all_data', methods=['POST'])
def clear_all_data():
    if 'username' not in session or session.get('role')!= 'admin':
        flash('Only admin can clear all data')
        return redirect('/dashboard')

    conn = get_db_connection()
    cur = conn.cursor()

    # Delete in order due to foreign key
    cur.execute("DELETE FROM drain_work_logs")
    cur.execute("DELETE FROM drains")

    conn.commit()
    cur.close()
    conn.close()
    flash('All drain data and work logs deleted. Users retained. You can re-import Excel now.')
    return redirect('/dashboard')

@app.route('/work_report')
def work_report():
    if 'username' not in session:
        return redirect('/login')

    if session.get('role')!= 'admin':
        flash('Access denied')
        return redirect('/dashboard')

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    cur.execute("""
        SELECT d.ward, d.drain_id, d.location, d.status as drain_status,
               wl.work_type, wl.status as work_status, wl.remarks,
               wl.uploaded_by, wl.uploaded_at, wl.photo_url
        FROM drains d
        LEFT JOIN drain_work_logs wl ON d.id = wl.drain_id
        ORDER BY d.ward, d.drain_id, wl.uploaded_at DESC
    """)

    work_data = cur.fetchall()
    cur.close()
    conn.close()

    return render_template('work_report.html', work_data=work_data)

@app.route('/download_excel')
def download_excel():
    if 'username' not in session or session.get('role')!= 'admin':
        flash('Only admin can download reports')
        return redirect('/dashboard')

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    cur.execute("""
        SELECT d.ward as "Ward",
               d.drain_id as "Drain ID",
               d.location as "Location",
               d.status as "Current Status",
               wl.work_type as "Work Type",
               wl.status as "Work Status",
               wl.remarks as "Remarks",
               wl.uploaded_by as "Updated By",
               wl.uploaded_at as "Work Date",
               wl.photo_url as "Photo URL"
        FROM drains d
        LEFT JOIN drain_work_logs wl ON d.id = wl.drain_id
        ORDER BY d.ward, d.drain_id, wl.uploaded_at DESC
    """)

    data = cur.fetchall()
    cur.close()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tanuku Drains Work Report"

    if data:
        headers = list(data[0].keys())
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        for row_num, row_data in enumerate(data, 2):
            for col_num, value in enumerate(row_data.values(), 1):
                if isinstance(value, datetime):
                    value = value.strftime('%d-%m-%Y %I:%M %p')
                ws.cell(row=row_num, column=col_num, value=value)

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Tanuku_Drains_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=filename)

@app.route('/logout')
def logout():
    session.clear()
    return redirect('/login')

if __name__ == '__main__':
    app.run(debug=False)
